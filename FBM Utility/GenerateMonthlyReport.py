import sys
import calendar
import datetime
from dateutil.relativedelta import relativedelta
import sys
import operator

import pandas as pd
import numpy as np
import openpyxl

from FoodBankManager import FBM
import FixedData


def add_data(chart, data, from_rows=False, titles_from_data=False, title=None):
	"""
	Add a range of data in a single pass.
	The default is to treat each column as a data series.
	"""
	if not isinstance(data, openpyxl.chart.Reference):
		data = openpyxl.chart.Reference(range_string=data)

	if from_rows:
		values = data.rows

	else:
		values = data.cols

	for v in values:
		range_string = u"{0}!{1}:{2}".format(data.sheetname, v[0], v[-1])
		series = openpyxl.chart.series_factory.SeriesFactory(range_string, title_from_data=titles_from_data, title=title)
		self.ser.append(series)

def FindLastMonthsDates():
	"""
	Finds the first and last days of last month
	:return: (first_of_month, last_of_month)
	:rtype: tuple(datetime.date, datetime.date)
	"""
	now = datetime.date.today()
	last_of_month = now.replace(day=1) - datetime.timedelta(days=1)
	first_of_month = last_of_month.replace(day=1)

	return first_of_month, last_of_month

def RunMonthlyReport(FBMInst,month=None, year=None):
	if month is not None and year is not None:
		start = datetime.date(year, month, 1)
		end = datetime.date(year, month, calendar.monthrange(year, month)[1])
	else:
		start, end = FindLastMonthsDates()

	data = FBMInst.GetFoodDonations(start, end)

	data[["Weight (lbs)"]] = data[["Weight (lbs)"]].astype("float")

	return data

def PivotInventoryTable(df):
	df[["Weight (lbs)"]] = df[["Weight (lbs)"]].astype("float")
	df = pd.pivot_table(df, index=["DonorCategory"], values=["Weight (lbs)"], aggfunc=[np.sum])
	return df

def MonthlyGuestData(FBMInst,month=None, year=None):
	if month is not None and year is not None:
		start = datetime.date(year, month, 1)
		end = datetime.date(year, month, calendar.monthrange(year, month)[1])
	else:
		start, end = FindLastMonthsDates()

	data = FBMInst.GetGuestData(start, end)

	data[["Tracking Result"]] = data[["Tracking Result"]].astype("int")

	return data

def WriteSummaryData(q, ws, origin=(1,1),  month=None, year=None, existing_clients=None, inventory_adjust=0, donor_catagories=[]):
	if month is not None and year is not None:
		start = datetime.date(year, month, 1)
		end = datetime.date(year, month, calendar.monthrange(year, month)[1])
	else:
		start, end = FindLastMonthsDates()

	donation_data = RunMonthlyReport(q, month=start.month, year=start.year)
	user_data = MonthlyGuestData(q, month=start.month, year=start.year)

	clients = set(user_data["Guest ID"].unique())
	if existing_clients is None:
		existing_clients = set()
	new_clients = set([c for c in clients if c not in existing_clients])

	ws.column_dimensions[ws.cell(row=origin[0], column=origin[1]).column].width = 14

	for cell_row in range(origin[0], origin[0] + len(donor_catagories) + 10 + 1):
		ws.cell(row=cell_row, column=origin[1]).style = "Input"

	ws.cell(row=origin[0], column=origin[1]).value = "{}".format(start.strftime("%B %Y"))
	ws.cell(row=origin[0], column=origin[1]).style = "Headline 4"
	ws.cell(row=origin[0], column=origin[1]).alignment = openpyxl.styles.Alignment(horizontal='center')
	for i, item in enumerate(donor_catagories):
		ws.cell(row=origin[0] + i + 1, column=origin[1]).value = donation_data[donation_data["DonorCategory"] == item].sum()["Weight (lbs)"]
	ws.cell(row=origin[0] + len(donor_catagories) + 1, column=origin[1]).value = "=SUM({}:{})".format(ws.cell(row=origin[0] + 1, column=origin[1]).coordinate, ws.cell(row=origin[0] + len(donor_catagories), column=origin[1]).coordinate)
	ws.cell(row=origin[0] + len(donor_catagories) + 1, column=origin[1]).style = "Calculation"
	ws.cell(row=origin[0] + len(donor_catagories) + 2, column=origin[1]).value = donation_data[donation_data["DonorCategory"] == "Waste"].sum()["Weight (lbs)"]
	ws.cell(row=origin[0] + len(donor_catagories) + 3, column=origin[1]).value = "={}-{}".format(ws.cell(row=origin[0] + len(donor_catagories) + 1, column=origin[1]).coordinate, ws.cell(row=origin[0] + len(donor_catagories) + 2, column=origin[1]).coordinate)
	ws.cell(row=origin[0] + len(donor_catagories) + 3, column=origin[1]).style = "Calculation"
	ws.cell(row=origin[0] + len(donor_catagories) + 4, column=origin[1]).style = "Normal"
	if (month, year) in FixedData.override.keys():
		ws.cell(row=origin[0] + len(donor_catagories) + 5, column=origin[1]).value = FixedData.override[(month, year)]["clients"]
                ws.cell(row=origin[0] + len(donor_catagories) + 6, column=origin[1]).value = FixedData.override[(month, year)]["new_clients"]
                ws.cell(row=origin[0] + len(donor_catagories) + 7, column=origin[1]).value = FixedData.override[(month, year)]["impact"]
	else:
		ws.cell(row=origin[0] + len(donor_catagories) + 5, column=origin[1]).value = len(clients)
		ws.cell(row=origin[0] + len(donor_catagories) + 6, column=origin[1]).value = len(new_clients)
		ws.cell(row=origin[0] + len(donor_catagories) + 7, column=origin[1]).value = user_data["Tracking Result"].sum()
	ws.cell(row=origin[0] + len(donor_catagories) + 8, column=origin[1]).value = "={}*{}".format(ws.cell(row=origin[0] + len(donor_catagories) + 5, column=origin[1]).coordinate, FixedData.output_weight)
	ws.cell(row=origin[0] + len(donor_catagories) + 9, column=origin[1]).style = "Normal"
	ws.cell(row=origin[0] + len(donor_catagories) + 10, column=origin[1]).value = "=IF(ISTEXT({0:}), {2:}-{1:}, {2:}-{1:}+{0:})+({3:})".format(ws.cell(row=origin[0] + len(donor_catagories) + 10, column=origin[1]-1).coordinate, ws.cell(row=origin[0] + len(donor_catagories) + 8, column=origin[1]).coordinate, ws.cell(row=origin[0] + len(donor_catagories) + 3, column=origin[1]).coordinate, inventory_adjust)
	
	return existing_clients.union(new_clients)


def WriteSummaryLabel(ws, origin=(1,1), donor_catagories=[]):
	ws.column_dimensions[ws.cell(row=origin[0], column=origin[1]).column].width = 22
	for i, item in enumerate(donor_catagories):
		ws.cell(row=origin[0] + i + 1, column=origin[1]).value = "{} (lbs)".format(item)
	ws.cell(row=origin[0] + len(donor_catagories) + 1, column=origin[1]).value = "Total Food Income (lbs)"
	ws.cell(row=origin[0] + len(donor_catagories) + 2, column=origin[1]).value = "Waste (lbs)"
	ws.cell(row=origin[0] + len(donor_catagories) + 3, column=origin[1]).value = "Total Collected (lbs)"

	ws.cell(row=origin[0] + len(donor_catagories) + 5, column=origin[1]).value = "Number of Clients"
	ws.cell(row=origin[0] + len(donor_catagories) + 6, column=origin[1]).value = "New Clients"
	ws.cell(row=origin[0] + len(donor_catagories) + 7, column=origin[1]).value = "Total Impact"
	ws.cell(row=origin[0] + len(donor_catagories) + 8, column=origin[1]).value = "Food Distributed (lbs)"

	ws.cell(row=origin[0] + len(donor_catagories) + 10, column=origin[1]).value = "Ending Inventory (lbs)"

	for cell_row in range(origin[0], origin[0] + len(donor_catagories) + 10 + 1):
		ws.cell(row=cell_row, column=origin[1]).style = "Headline 4"

def WriteExcelSheet(name, month=None, year=None):
	if month is not None and year is not None:
		start = datetime.date(year, month, 1)
		end = datetime.date(year, month, calendar.monthrange(year, month)[1])
	else:
		start, end = FindLastMonthsDates()

	q = FBM("mcfb.soxbox.co")

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "12 Month Overview"
	
	# borrowed code from GenerateGraphs.py figures inventory adjustment and proper categories
	if month is not None and year is not None:
		x_end = datetime.datetime(year, month, 1)
	else:
		x_now = datetime.datetime.now()
		x_end = datetime.datetime(x_now.year, x_now.month, 1)
	x_end = x_end + relativedelta(months=+1)
	x_start = x_end + relativedelta(months=-13)
	food_data = q.GetFoodDonations(x_start, x_end)
	guest_data = q.GetGuestData(x_start, x_end)
	food_data[u'Donated On'] = pd.to_datetime(food_data[u'Donated On']).astype(datetime.datetime)
	food_data[u'Weight (lbs)'] = food_data[u'Weight (lbs)'].astype(float)
	guest_data[u'Outreach on'] = pd.to_datetime(guest_data[u'Outreach on']).astype(datetime.datetime)
	guest_data[u'Tracking Result'] = guest_data[u'Tracking Result'].astype(int)
	
	period_start = x_start
	inventory = 0
	inventory_adjust = 0
	categories = set(["Grocery", "Org/Corp"])
	
	for i in range(13):
		period_end = period_start + relativedelta(months=+1)
		
		food_month = food_data[(food_data[u'Donated On'] >= period_start) & (food_data[u'Donated On'] < period_end)]
		guest_month = guest_data[(guest_data[u'Outreach on'] >= period_start) & (guest_data[u'Outreach on'] < period_end)]
		
		month_clients = len(set(guest_month[u'Guest ID'].unique()))
		
		if (period_start.month, period_start.year) in FixedData.override.keys():
			month_clients = FixedData.override[(period_start.month, period_start.year)]['clients']
		
		intake_total = food_month[food_month[u'DonorCategory'] != u'Waste'][u'Weight (lbs)'].sum()
		waste_total =  food_month[food_month[u'DonorCategory'] == u'Waste'][u'Weight (lbs)'].sum()
		output_total = month_clients * FixedData.output_weight
		food_out_total = waste_total + output_total
		
		if i > 0 and (period_start.month, period_start.year) in FixedData.inventory.keys():
			inventory_adjust = (FixedData.inventory[(period_start.month, period_start.year)] - (intake_total - food_out_total)) - inventory
			#print("FixedData.inventory[({}, {})] = {}; intake_total = {}; waste_total = {}; output_total = {}; food_out_total = {}; month net food is {}; tallied inventory is {} => inventory_adjust = {}".format(period_start.month, period_start.year, FixedData.inventory[(period_start.month, period_start.year)], intake_total, waste_total, output_total, food_out_total, intake_total - food_out_total, inventory, inventory_adjust))
		
		if i > 0:
			inventory += (intake_total - (waste_total + output_total))
			#print("intake={}, waste={}, output={}, delta={}, inventory={}".format(intake_total, waste_total, output_total, (intake_total - (waste_total + output_total)), inventory))
		
		food_datums = RunMonthlyReport(q, month=period_start.month, year=period_start.year)
		
		if not food_datums.empty:
			messy_data = PivotInventoryTable(food_datums)
			useful_data = messy_data.to_dict()[('sum', 'Weight (lbs)')]
			sorted_data = zip(*sorted(useful_data.items(), key=operator.itemgetter(1), reverse=True))
			categories = categories.union(sorted_data[0])
		
		period_start = period_end
	
	if u'Waste' in categories:
		categories.remove(u'Waste')
	
	WriteSummaryLabel(ws, origin=(2, 1), donor_catagories=categories)
	ym_start = 12 * year + month - 1
	clients = set()
	adjusted = False
	for i, ym in enumerate(range(ym_start - 11, ym_start + 1)):
		y, m = divmod(ym, 12)
		if not adjusted:
			clients = WriteSummaryData(q, ws, origin=(2, i+2), month=m+1, year=y, existing_clients=clients, inventory_adjust=inventory_adjust, donor_catagories=categories)
			adjusted = True
		else:
			clients = WriteSummaryData(q, ws, origin=(2, i+2), month=m+1, year=y, existing_clients=clients, donor_catagories=categories)

	ws.merge_cells('B1:M1')
	for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
		ws['{}1'.format(col)].style = "Headline 1"
	ws['B1'] = "12 Month Overview"
	ws['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')

	ws['N1'] = "Last Year's Performance"
	ws['N1'].style = "Headline 1"
	WriteSummaryData(q, ws, origin=(2, 14), month=month, year=year-1, donor_catagories=categories)
	ws.column_dimensions['N'].width = 30
	ws['N{}'.format(8+len(categories))].style = "Normal"
	ws['N{}'.format(8+len(categories))] = ""
	ws['N{}'.format(12+len(categories))].style = "Normal"
	ws['N{}'.format(12+len(categories))] = ""

	ws.freeze_panes = "B3"

	c1 = openpyxl.chart.LineChart()
	c1.title = "Food Income (Lbs), Large sources"
	data = openpyxl.chart.Reference(*(ws,) + openpyxl.utils.cell.range_boundaries("A3:M4"))
	c1.add_data(data, titles_from_data=True, from_rows=True)
	ws.add_chart(c1, "B22")

	c1 = openpyxl.chart.LineChart()
	c1.title = "Food Income (Lbs), All Other Sources"
	data = openpyxl.chart.Reference(*(ws,) + openpyxl.utils.cell.range_boundaries("A5:M{}".format(2+len(categories))))
	c1.add_data(data, titles_from_data=True, from_rows=True)
	ws.add_chart(c1, "H22")

	c1 = openpyxl.chart.LineChart()
	c1.title = "Ending Inventory (lbs)"
	data = openpyxl.chart.Reference(*(ws,) + openpyxl.utils.cell.range_boundaries("A{0}:M{0}".format(12+len(categories))))
	c1.add_data(data, titles_from_data=True, from_rows=True)
	ws.add_chart(c1, "N22")

	c1 = openpyxl.chart.LineChart()
	c1.title = "Waste (lbs)"
	data = openpyxl.chart.Reference(*(ws,) + openpyxl.utils.cell.range_boundaries("A{0}:M{0}".format(4+len(categories))))
	c1.add_data(data, titles_from_data=True, from_rows=True)
	ws.add_chart(c1, "B37")

	c1 = openpyxl.chart.LineChart()
	c1.title = "Clients"
	data = openpyxl.chart.Reference(*(ws,) + openpyxl.utils.cell.range_boundaries("A{0}:M{1}".format(7+len(categories), 9+len(categories))))
	c1.add_data(data, titles_from_data=True, from_rows=True)
	ws.add_chart(c1, "H37")

	filename = "{}.xlsx".format(name)
	wb.save(filename)
	return filename

if __name__ == '__main__':
	pd.set_option('display.expand_frame_repr', False)
	if len(sys.argv) < 3:
		print "Run with \"<month number (1-12)> <Year (4 digit)>\""
	print WriteExcelSheet("out/Report {}-{}".format(sys.argv[1], sys.argv[2]), month=int(sys.argv[1]), year=int(sys.argv[2]))
